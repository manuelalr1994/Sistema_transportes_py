import os
import datetime
from operator import truediv
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font

from django.shortcuts import render, redirect
from django.http import JsonResponse, QueryDict, HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import View
from django.template.loader import get_template

from applications.fletes.catalogos.camiones.models import Camiones
from applications.fletes.catalogos.compania_fletera.models import CompaniasFleteras
from applications.nomina.catalogos.campos_agricolas.models import NomCampos
from applications.nomina.catalogos.cuadrilleros.models import NomCuadrilleros
from applications.nomina.catalogos.cultivos.models import Cultivos, Variedad
from applications.nomina.catalogos.labores.models import NomLabores
from applications.nomina.catalogos.semanas.models import NomSemanas, NomTipoSemanas
from applications.fletes.procesos.capturas_fleteras.models import FleJornada
from applications.fletes.procesos.listados_nomina.forms import FormularioListadoNomina

# Create your views here.

def listadosNomina(request):
    if request.user.permisos_fletes:
        template_name = "fletes/procesos/listados_nomina/listados_nomina.html"
        data = {}
        context = {}
        context['form'] = FormularioListadoNomina
        context['dias_semana'] = []

        if request.method == "POST":
            return render(request, template_name, context)
        else:  
            return render(request, template_name, context)
    else:
        return redirect('users_app:user_error')


def cargarListaNomina(request):
    data = {}
    diccionario_dias = {0:'lunes', 1:'martes', 2:'miercoles', 3:'jueves', 4:'viernes', 5:'sabado', 6:'domingo'}

    empresa = int(request.GET.get('empresa'))
    tipo_semana = int(request.GET.get('tipo_semana'))
    semana = int(request.GET.get('semana'))

    lista_camiones = FleJornada.objects.filter(campo_agricola=empresa, tipo_semana=tipo_semana, semana=semana, cerrada=True).values_list('camion', flat=True).distinct()
    lista_capturas = FleJornada.objects.filter(campo_agricola=empresa, tipo_semana=tipo_semana, semana=semana, cerrada=True).order_by('fecha')
    lista_capturas_final = []
    dict_captura = {}
    horas_totales = 0
    total_importe = 0

    # Con la lista de camiones extraida, creamos diccionarios para cada camion
    # con la informacion de las capturas de cada uno para esa semana
    for camion in lista_camiones:

        horas_totales = 0
        total_importe = 0

        # Inicializamos a 0 las horas trabajadas en todos los dias de la semana
        for dia in diccionario_dias.values():
            dict_captura[dia] = 0

        # Recorremos las capturas del camion actual y extraemos 
        # sus horas trabajadas y las horas por dia
        for captura in lista_capturas:
            if captura.camion.pk == camion:
                dict_captura['id'] = captura.camion.pk
                dict_captura['codigo'] = captura.camion.codigo
                dict_captura['nombre'] = captura.camion.nombre
                dia_semana_captura = diccionario_dias[captura.fecha.weekday()]
                dict_captura[dia_semana_captura] = captura.total_hrs
                horas_totales = round(horas_totales + captura.total_hrs, 2)
                total_importe = round((total_importe + (captura.camion.costo_hra * captura.total_hrs)), 2)

        dict_captura['horas'] = horas_totales
        dict_captura['sub_total'] = total_importe
        dict_captura['diesel'] = 0.00
        dict_captura['comida'] = 0.00
        dict_captura['total'] = total_importe

        lista_capturas_final.append(dict_captura.copy())

    data['lista_capturas'] = lista_capturas_final
    return render(request, 'fletes/procesos/listados_nomina/cargar_listado_nomina.html', data)


class ReporteListadoCamiones(View):

    def get(self, request, *args, **kwargs):

        try:
            context = {}

            # Generamos el Contexto para el Template
            diccionario_dias = {0:'lunes', 1:'martes', 2:'miercoles', 3:'jueves', 4:'viernes', 5:'sabado', 6:'domingo'}

            empresa = int(request.GET.get('empresa'))
            tipo_semana = int(request.GET.get('tipo_semana'))
            semana = int(request.GET.get('semana'))

            # Definimos las fecha, hora y numero de la cabecera
            semana_query = NomSemanas.objects.get(pk=semana)
            context["numero_semana"] = semana_query.semana
            context["inicio"] = semana_query.fecha_inicial.strftime("%d/%m/%y")
            context["cierre"] = semana_query.fecha_final.strftime("%d/%m/%y")
            context["dia_reporte"] = datetime.datetime.now().strftime("%d/%m/%y")
            context["hora_reporte"] = datetime.datetime.now().strftime("%H:%M:%S")


            lista_camiones = FleJornada.objects.filter(campo_agricola=empresa, tipo_semana=tipo_semana, semana=semana, cerrada=True).values_list('camion', flat=True).distinct()
            lista_capturas = FleJornada.objects.filter(campo_agricola=empresa, tipo_semana=tipo_semana, semana=semana, cerrada=True).order_by('fecha')
            lista_capturas_final = []
            dict_captura = {}
            total_general = 0 # Este llevara la suma de los totales de importe de todos los camiones
            total_horas_general = 0 # Este llevara la suma de los totales de horas de todos los camiones
            horas_totales = 0 # Este llevara el total de horas del camion
            total_importe = 0 # Este llevara el importe total del camion

            # Con la lista de camiones extraida, creamos diccionarios para cada camion
            # con la informacion de las capturas de cada uno para esa semana
            for camion in lista_camiones:

                horas_totales = 0
                total_importe = 0

                # Inicializamos a 0 las horas trabajadas en todos los dias de la semana
                for dia in diccionario_dias.values():
                    dict_captura[dia] = 0

                # Recorremos las capturas del camion actual y extraemos 
                # sus horas trabajadas y las horas por dia
                for captura in lista_capturas:
                    if captura.camion.pk == camion:
                        dict_captura['id'] = captura.camion.pk
                        dict_captura['codigo'] = str(captura.camion.codigo).rjust(5, '0')
                        dict_captura['nombre'] = captura.camion.nombre
                        dia_semana_captura = diccionario_dias[captura.fecha.weekday()]
                        dict_captura[dia_semana_captura] = captura.total_hrs
                        horas_totales = round(horas_totales + captura.total_hrs, 2)
                        total_importe = round((total_importe + (captura.camion.costo_hra * captura.total_hrs)), 2)

                dict_captura['horas'] = horas_totales
                dict_captura['sub_total'] = total_importe
                dict_captura['diesel'] = round(0, 2)
                dict_captura['comida'] = round(0, 2)
                dict_captura['total'] = total_importe
                total_general = total_general + total_importe
                total_horas_general = total_horas_general + horas_totales

                lista_capturas_final.append(dict_captura.copy())

            context['lista_capturas'] = lista_capturas_final
            context['nombre_reporte'] = "LISTADO GENERAL DE CAMIONES"
            context['total_general'] = round(total_general, 2)
            context['total_horas_general'] = round(total_horas_general, 2)
            
            template = get_template("fletes/procesos/listados_nomina/reportes/reporte_listado_camiones.html")
            html_template = template.render(context)
            pdf = HTML(string=html_template).write_pdf()
            print('Estoy tratando de enviar el HTTP RESPONSE PDF')
            return HttpResponse(pdf, content_type='application/pdf')
            
        except:
            pass

        return HttpResponseRedirect(reverse_lazy('fletes:procesos:listados_nomina:lista'))


class ReporteExcel(View):
    def get(self, request, *args, **kwargs):

        campo_agricola = request.GET.get('campo_agricola')
        tipo_semana = request.GET.get('tipo_semana')
        semana = request.GET.get('semana')

        # Recabamos informacion y la organizamos para el diccionario
        lista_capturas = FleJornada.objects.filter(campo_agricola=campo_agricola, tipo_semana=tipo_semana, semana=semana)
        depositos_por_campo = {}

        for captura in lista_capturas:
            if captura.camion.compania_fletera.pk in depositos_por_campo.keys():
                print(depositos_por_campo[captura.camion.compania_fletera.pk]['total_depositar'])
                depositos_por_campo[captura.camion.compania_fletera.pk]['total_depositar'] += int(captura.importe)
            else:
                depositos_por_campo[captura.camion.compania_fletera.pk] = {}
                depositos_por_campo[captura.camion.compania_fletera.pk]['codigo'] = captura.camion.compania_fletera.codigo
                depositos_por_campo[captura.camion.compania_fletera.pk]['nombre'] = captura.camion.compania_fletera.nombre
                depositos_por_campo[captura.camion.compania_fletera.pk]['total_depositar'] = int(captura.importe )
                

        # Desarrollamos el Excel
        wb = Workbook()
        ws = wb.active
        font_titulo = Font(bold=True)

        # Especificamos el Titulo (campo agricola)
        ws['A1'] = NomCampos.objects.get(pk=campo_agricola).nombre
        ws['A1'].font = font_titulo
        ws.merge_cells('A1:B1')

        # Especificamos la Semana
        fecha_inicial =  NomSemanas.objects.get(pk=semana, tipo_semana=tipo_semana).fecha_inicial
        fecha_final =  NomSemanas.objects.get(pk=semana, tipo_semana=tipo_semana).fecha_final
        ws['A2'] = "Semana del "+fecha_inicial.strftime("%d/%m/%Y")+" a "+fecha_final.strftime("%d/%m/%Y")
        ws.merge_cells('A2:B2')

        # Especificamos la cabecera de la tabla
        ws['A4'] = 'CODIGO'
        ws['A4'].font = font_titulo
        ws['B4'] = 'NOMBRE'
        ws['B4'].font = font_titulo
        ws['C4'] = 'TOT DEPOSITAR'
        ws['C4'].font = font_titulo
        ws['D4'] = 'NO CUENTA'
        ws['D4'].font = font_titulo
        ws['E4'] = 'NO TARJETA'
        ws['E4'].font = font_titulo

        cont = 5

        for campo in depositos_por_campo:
            ws.cell(row = cont, column = 1).value = depositos_por_campo[campo]['codigo']
            ws.cell(row = cont, column = 2).value = depositos_por_campo[campo]['nombre']
            ws.cell(row = cont, column = 3).value = depositos_por_campo[campo]['total_depositar']
            cont += 1

        # Ajustamos las dimensiones de las columnas
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type= "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response